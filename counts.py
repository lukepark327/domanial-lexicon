import os
from openpyxl import Workbook


def news_count(IN_PATH="./result"):
    ret = {}

    keywords = os.listdir(IN_PATH)
    for keyword in keywords:
        years = os.listdir(IN_PATH + '/' + keyword)
        for year in years:
            cnt = 0

            days = os.listdir(IN_PATH + '/' + keyword + '/' + year)
            for day in days:
                tars = os.listdir(IN_PATH + '/' + keyword + '/' + year + '/' + day)
                for tar in tars:

                    # file open
                    with open(IN_PATH + '/' + keyword + '/' + year + '/' + day + '/' + tar, encoding='UTF-8') as rf:
                        lines = rf.readlines()
                        cnt += len(lines) // 3

            # print(keyword, year, cnt)
            ret[keyword+year] = cnt

    return ret


def word_count(IN_PATH="./corpus", keywords=['불행', '행복']):
    ret = {}

    years = os.listdir(IN_PATH)
    for year in years:
        for keyword in keywords:
            with open(IN_PATH + '/' + year + '/' + keyword + '.txt', encoding='UTF-8') as rf:
                words = rf.read().strip().split()

                # print(year, keyword, len(words), len(set(words)))
                ret[keyword + year] = (len(words), len(set(words)))

    return ret


if __name__ == "__main__":
    OUT_PATH = "./count"
    if not os.path.exists(OUT_PATH):
        os.makedirs(OUT_PATH)

    """
    processing
    """
    n_cnt = news_count(IN_PATH="./result")
    w_cnt = word_count(IN_PATH="./corpus", keywords=['불행', '행복'])

    """
    save excel
    """
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='')
    ws.cell(row=1, column=2, value='news')
    ws.cell(row=1, column=3, value='words')
    ws.cell(row=1, column=4, value='deduplication')

    years = list(range(2000, 2018))  # years range: 2000 ~ 2017
    keywords = ['불행', '행복']

    idx = 2

    for keyword in keywords:
        for year in years:
            ws.cell(row=idx, column=1, value=keyword + '_' + str(year))
            ws.cell(row=idx, column=2, value=n_cnt[keyword + str(year)])  # num of news
            ws.cell(row=idx, column=3, value=w_cnt[keyword + str(year)][0])  # num of words
            ws.cell(row=idx, column=4, value=w_cnt[keyword + str(year)][1])  # num of words - deduplication

            idx += 1

    wb.save(OUT_PATH + '/' + "count.xlsx")
